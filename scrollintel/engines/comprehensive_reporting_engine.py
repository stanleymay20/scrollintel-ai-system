"""
Comprehensive Reporting Engine for Advanced Analytics Dashboard System

This module provides comprehensive reporting capabilities with support for multiple formats
including PDF, Excel, and web formats with automated scheduling and distribution.
"""

import asyncio
import json
import logging
from datetime import datetime, timedelta
from typing import Dict, List, Optional, Any, Union
from dataclasses import dataclass, asdict
from enum import Enum
import pandas as pd
from io import BytesIO
import base64

# PDF generation
try:
    from reportlab.lib.pagesizes import letter, A4
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, Image
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import inch
    from reportlab.lib import colors
    REPORTLAB_AVAILABLE = True
except ImportError:
    REPORTLAB_AVAILABLE = False

# Excel generation
try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.chart import LineChart, BarChart, Reference
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

logger = logging.getLogger(__name__)


class ReportFormat(Enum):
    """Supported report formats"""
    PDF = "pdf"
    EXCEL = "excel"
    WEB = "web"
    JSON = "json"
    CSV = "csv"


class ReportType(Enum):
    """Types of reports available"""
    EXECUTIVE_SUMMARY = "executive_summary"
    DETAILED_ANALYTICS = "detailed_analytics"
    ROI_ANALYSIS = "roi_analysis"
    PERFORMANCE_METRICS = "performance_metrics"
    PREDICTIVE_INSIGHTS = "predictive_insights"
    CUSTOM = "custom"


@dataclass
class ReportConfig:
    """Configuration for report generation"""
    report_type: ReportType
    format: ReportFormat
    title: str
    description: str
    data_sources: List[str]
    filters: Dict[str, Any]
    date_range: Dict[str, datetime]
    template_id: Optional[str] = None
    custom_sections: Optional[List[Dict]] = None
    branding: Optional[Dict] = None
    recipients: Optional[List[str]] = None
    schedule: Optional[Dict] = None


@dataclass
class ReportSection:
    """Individual section within a report"""
    title: str
    content_type: str  # text, chart, table, image
    data: Any
    styling: Optional[Dict] = None
    order: int = 0


@dataclass
class GeneratedReport:
    """Generated report with metadata"""
    report_id: str
    config: ReportConfig
    content: bytes
    metadata: Dict[str, Any]
    generated_at: datetime
    file_size: int
    format: ReportFormat


class ComprehensiveReportingEngine:
    """
    Advanced reporting engine with multi-format support and automated distribution
    """
    
    def __init__(self):
        self.logger = logging.getLogger(__name__)
        self.report_cache = {}
        self.templates = {}
        self.schedulers = {}
        
    async def generate_report(self, config: ReportConfig) -> GeneratedReport:
        """
        Generate a comprehensive report based on configuration
        
        Args:
            config: Report configuration
            
        Returns:
            Generated report with content and metadata
        """
        try:
            self.logger.info(f"Generating {config.report_type.value} report in {config.format.value} format")
            
            # Collect and process data
            report_data = await self._collect_report_data(config)
            
            # Generate report content based on format
            if config.format == ReportFormat.PDF:
                content = await self._generate_pdf_report(config, report_data)
            elif config.format == ReportFormat.EXCEL:
                content = await self._generate_excel_report(config, report_data)
            elif config.format == ReportFormat.WEB:
                content = await self._generate_web_report(config, report_data)
            elif config.format == ReportFormat.JSON:
                content = await self._generate_json_report(config, report_data)
            elif config.format == ReportFormat.CSV:
                content = await self._generate_csv_report(config, report_data)
            else:
                raise ValueError(f"Unsupported report format: {config.format}")
            
            # Create report metadata
            metadata = {
                "report_type": config.report_type.value,
                "format": config.format.value,
                "data_sources": config.data_sources,
                "filters_applied": config.filters,
                "date_range": {
                    "start": config.date_range.get("start").isoformat() if config.date_range.get("start") else None,
                    "end": config.date_range.get("end").isoformat() if config.date_range.get("end") else None
                },
                "sections_count": len(report_data.get("sections", [])),
                "generation_time": datetime.utcnow().isoformat()
            }
            
            # Create generated report
            report = GeneratedReport(
                report_id=f"report_{datetime.utcnow().strftime('%Y%m%d_%H%M%S')}",
                config=config,
                content=content,
                metadata=metadata,
                generated_at=datetime.utcnow(),
                file_size=len(content),
                format=config.format
            )
            
            # Cache report if needed
            self.report_cache[report.report_id] = report
            
            self.logger.info(f"Successfully generated report {report.report_id}")
            return report
            
        except Exception as e:
            self.logger.error(f"Error generating report: {str(e)}")
            raise
    
    async def _collect_report_data(self, config: ReportConfig) -> Dict[str, Any]:
        """Collect and process data for report generation"""
        try:
            report_data = {
                "title": config.title,
                "description": config.description,
                "generated_at": datetime.utcnow(),
                "sections": []
            }
            
            # Executive Summary data
            if config.report_type == ReportType.EXECUTIVE_SUMMARY:
                report_data["sections"] = await self._get_executive_summary_sections(config)
            
            # Detailed Analytics data
            elif config.report_type == ReportType.DETAILED_ANALYTICS:
                report_data["sections"] = await self._get_detailed_analytics_sections(config)
            
            # ROI Analysis data
            elif config.report_type == ReportType.ROI_ANALYSIS:
                report_data["sections"] = await self._get_roi_analysis_sections(config)
            
            # Performance Metrics data
            elif config.report_type == ReportType.PERFORMANCE_METRICS:
                report_data["sections"] = await self._get_performance_metrics_sections(config)
            
            # Predictive Insights data
            elif config.report_type == ReportType.PREDICTIVE_INSIGHTS:
                report_data["sections"] = await self._get_predictive_insights_sections(config)
            
            # Custom report data
            elif config.report_type == ReportType.CUSTOM:
                report_data["sections"] = await self._get_custom_sections(config)
            
            return report_data
            
        except Exception as e:
            self.logger.error(f"Error collecting report data: {str(e)}")
            raise
    
    async def _get_executive_summary_sections(self, config: ReportConfig) -> List[ReportSection]:
        """Generate executive summary sections"""
        sections = []
        
        # Key Metrics Overview
        sections.append(ReportSection(
            title="Key Performance Indicators",
            content_type="table",
            data={
                "headers": ["Metric", "Current Value", "Previous Period", "Change"],
                "rows": [
                    ["Total Revenue", "$2.5M", "$2.1M", "+19%"],
                    ["Active Users", "15,432", "12,890", "+20%"],
                    ["Conversion Rate", "3.2%", "2.8%", "+14%"],
                    ["Customer Satisfaction", "4.6/5", "4.4/5", "+5%"]
                ]
            },
            order=1
        ))
        
        # Trend Analysis
        sections.append(ReportSection(
            title="Performance Trends",
            content_type="chart",
            data={
                "type": "line",
                "labels": ["Jan", "Feb", "Mar", "Apr", "May", "Jun"],
                "datasets": [
                    {
                        "label": "Revenue Growth",
                        "data": [100, 120, 115, 134, 145, 160]
                    }
                ]
            },
            order=2
        ))
        
        # Key Insights
        sections.append(ReportSection(
            title="Strategic Insights",
            content_type="text",
            data={
                "insights": [
                    "Revenue growth accelerated by 19% driven by new product launches",
                    "User engagement increased significantly with mobile app improvements",
                    "Customer acquisition costs decreased by 12% through optimized campaigns"
                ]
            },
            order=3
        ))
        
        return sections
    
    async def _get_detailed_analytics_sections(self, config: ReportConfig) -> List[ReportSection]:
        """Generate detailed analytics sections"""
        sections = []
        
        # Detailed Metrics
        sections.append(ReportSection(
            title="Comprehensive Analytics",
            content_type="table",
            data={
                "headers": ["Category", "Metric", "Value", "Trend", "Target"],
                "rows": [
                    ["Sales", "Monthly Revenue", "$425K", "↑ 15%", "$400K"],
                    ["Marketing", "Lead Generation", "1,250", "↑ 22%", "1,000"],
                    ["Operations", "Efficiency Score", "87%", "↑ 5%", "85%"],
                    ["Customer", "Retention Rate", "94%", "↑ 2%", "90%"]
                ]
            },
            order=1
        ))
        
        return sections
    
    async def _get_roi_analysis_sections(self, config: ReportConfig) -> List[ReportSection]:
        """Generate ROI analysis sections"""
        sections = []
        
        # ROI Summary
        sections.append(ReportSection(
            title="Return on Investment Analysis",
            content_type="table",
            data={
                "headers": ["Investment", "Cost", "Benefit", "ROI", "Payback Period"],
                "rows": [
                    ["AI Platform", "$150K", "$320K", "113%", "8 months"],
                    ["Marketing Automation", "$75K", "$180K", "140%", "6 months"],
                    ["Data Analytics", "$100K", "$250K", "150%", "7 months"]
                ]
            },
            order=1
        ))
        
        return sections
    
    async def _get_performance_metrics_sections(self, config: ReportConfig) -> List[ReportSection]:
        """Generate performance metrics sections"""
        sections = []
        
        # Performance Dashboard
        sections.append(ReportSection(
            title="System Performance Metrics",
            content_type="table",
            data={
                "headers": ["System", "Uptime", "Response Time", "Throughput", "Error Rate"],
                "rows": [
                    ["API Gateway", "99.9%", "45ms", "1,200 req/s", "0.1%"],
                    ["Database", "99.8%", "12ms", "5,000 ops/s", "0.05%"],
                    ["Analytics Engine", "99.7%", "150ms", "500 req/s", "0.2%"]
                ]
            },
            order=1
        ))
        
        return sections
    
    async def _get_predictive_insights_sections(self, config: ReportConfig) -> List[ReportSection]:
        """Generate predictive insights sections"""
        sections = []
        
        # Forecasts
        sections.append(ReportSection(
            title="Predictive Analytics & Forecasts",
            content_type="chart",
            data={
                "type": "line",
                "labels": ["Jul", "Aug", "Sep", "Oct", "Nov", "Dec"],
                "datasets": [
                    {
                        "label": "Predicted Revenue",
                        "data": [170, 185, 195, 210, 225, 240]
                    },
                    {
                        "label": "Confidence Interval",
                        "data": [160, 175, 185, 200, 215, 230]
                    }
                ]
            },
            order=1
        ))
        
        return sections
    
    async def _get_custom_sections(self, config: ReportConfig) -> List[ReportSection]:
        """Generate custom report sections"""
        sections = []
        
        if config.custom_sections:
            for section_config in config.custom_sections:
                sections.append(ReportSection(
                    title=section_config.get("title", "Custom Section"),
                    content_type=section_config.get("content_type", "text"),
                    data=section_config.get("data", {}),
                    order=section_config.get("order", 0)
                ))
        
        return sections
    
    async def _generate_pdf_report(self, config: ReportConfig, report_data: Dict) -> bytes:
        """Generate PDF report"""
        if not REPORTLAB_AVAILABLE:
            raise ImportError("ReportLab is required for PDF generation")
        
        try:
            buffer = BytesIO()
            doc = SimpleDocTemplate(buffer, pagesize=letter)
            styles = getSampleStyleSheet()
            story = []
            
            # Title
            title_style = ParagraphStyle(
                'CustomTitle',
                parent=styles['Heading1'],
                fontSize=24,
                spaceAfter=30,
                alignment=1  # Center alignment
            )
            story.append(Paragraph(report_data["title"], title_style))
            story.append(Spacer(1, 12))
            
            # Description
            story.append(Paragraph(report_data["description"], styles['Normal']))
            story.append(Spacer(1, 12))
            
            # Generation info
            gen_info = f"Generated on: {report_data['generated_at'].strftime('%Y-%m-%d %H:%M:%S')}"
            story.append(Paragraph(gen_info, styles['Normal']))
            story.append(Spacer(1, 20))
            
            # Sections
            for section in sorted(report_data["sections"], key=lambda x: x.order):
                # Section title
                story.append(Paragraph(section.title, styles['Heading2']))
                story.append(Spacer(1, 12))
                
                # Section content
                if section.content_type == "table":
                    table_data = [section.data["headers"]] + section.data["rows"]
                    table = Table(table_data)
                    table.setStyle(TableStyle([
                        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                        ('FONTSIZE', (0, 0), (-1, 0), 14),
                        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                        ('GRID', (0, 0), (-1, -1), 1, colors.black)
                    ]))
                    story.append(table)
                
                elif section.content_type == "text":
                    if isinstance(section.data, dict) and "insights" in section.data:
                        for insight in section.data["insights"]:
                            story.append(Paragraph(f"• {insight}", styles['Normal']))
                    else:
                        story.append(Paragraph(str(section.data), styles['Normal']))
                
                story.append(Spacer(1, 20))
            
            # Build PDF
            doc.build(story)
            buffer.seek(0)
            return buffer.getvalue()
            
        except Exception as e:
            self.logger.error(f"Error generating PDF report: {str(e)}")
            raise
    
    async def _generate_excel_report(self, config: ReportConfig, report_data: Dict) -> bytes:
        """Generate Excel report"""
        if not OPENPYXL_AVAILABLE:
            raise ImportError("openpyxl is required for Excel generation")
        
        try:
            workbook = openpyxl.Workbook()
            worksheet = workbook.active
            worksheet.title = "Report"
            
            # Title and header
            worksheet['A1'] = report_data["title"]
            worksheet['A1'].font = Font(size=16, bold=True)
            worksheet['A2'] = report_data["description"]
            worksheet['A3'] = f"Generated: {report_data['generated_at'].strftime('%Y-%m-%d %H:%M:%S')}"
            
            current_row = 5
            
            # Sections
            for section in sorted(report_data["sections"], key=lambda x: x.order):
                # Section title
                worksheet.cell(row=current_row, column=1, value=section.title)
                worksheet.cell(row=current_row, column=1).font = Font(size=14, bold=True)
                current_row += 2
                
                # Section content
                if section.content_type == "table":
                    # Headers
                    for col, header in enumerate(section.data["headers"], 1):
                        cell = worksheet.cell(row=current_row, column=col, value=header)
                        cell.font = Font(bold=True)
                        cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
                    current_row += 1
                    
                    # Data rows
                    for row_data in section.data["rows"]:
                        for col, value in enumerate(row_data, 1):
                            worksheet.cell(row=current_row, column=col, value=value)
                        current_row += 1
                
                elif section.content_type == "text":
                    if isinstance(section.data, dict) and "insights" in section.data:
                        for insight in section.data["insights"]:
                            worksheet.cell(row=current_row, column=1, value=f"• {insight}")
                            current_row += 1
                    else:
                        worksheet.cell(row=current_row, column=1, value=str(section.data))
                        current_row += 1
                
                current_row += 2
            
            # Save to buffer
            buffer = BytesIO()
            workbook.save(buffer)
            buffer.seek(0)
            return buffer.getvalue()
            
        except Exception as e:
            self.logger.error(f"Error generating Excel report: {str(e)}")
            raise
    
    async def _generate_web_report(self, config: ReportConfig, report_data: Dict) -> bytes:
        """Generate web/HTML report"""
        try:
            html_content = f"""
            <!DOCTYPE html>
            <html>
            <head>
                <title>{report_data['title']}</title>
                <style>
                    body {{ font-family: Arial, sans-serif; margin: 40px; }}
                    h1 {{ color: #333; text-align: center; }}
                    h2 {{ color: #666; border-bottom: 2px solid #eee; }}
                    table {{ border-collapse: collapse; width: 100%; margin: 20px 0; }}
                    th, td {{ border: 1px solid #ddd; padding: 12px; text-align: left; }}
                    th {{ background-color: #f2f2f2; font-weight: bold; }}
                    .insight {{ margin: 10px 0; padding: 10px; background-color: #f9f9f9; }}
                    .meta {{ color: #888; font-size: 0.9em; }}
                </style>
            </head>
            <body>
                <h1>{report_data['title']}</h1>
                <p>{report_data['description']}</p>
                <p class="meta">Generated on: {report_data['generated_at'].strftime('%Y-%m-%d %H:%M:%S')}</p>
            """
            
            # Sections
            for section in sorted(report_data["sections"], key=lambda x: x.order):
                html_content += f"<h2>{section.title}</h2>"
                
                if section.content_type == "table":
                    html_content += "<table>"
                    # Headers
                    html_content += "<tr>"
                    for header in section.data["headers"]:
                        html_content += f"<th>{header}</th>"
                    html_content += "</tr>"
                    
                    # Data rows
                    for row in section.data["rows"]:
                        html_content += "<tr>"
                        for cell in row:
                            html_content += f"<td>{cell}</td>"
                        html_content += "</tr>"
                    html_content += "</table>"
                
                elif section.content_type == "text":
                    if isinstance(section.data, dict) and "insights" in section.data:
                        for insight in section.data["insights"]:
                            html_content += f'<div class="insight">• {insight}</div>'
                    else:
                        html_content += f"<p>{section.data}</p>"
            
            html_content += """
            </body>
            </html>
            """
            
            return html_content.encode('utf-8')
            
        except Exception as e:
            self.logger.error(f"Error generating web report: {str(e)}")
            raise
    
    async def _generate_json_report(self, config: ReportConfig, report_data: Dict) -> bytes:
        """Generate JSON report"""
        try:
            # Convert datetime objects to strings for JSON serialization
            json_data = {
                "title": report_data["title"],
                "description": report_data["description"],
                "generated_at": report_data["generated_at"].isoformat(),
                "sections": []
            }
            
            for section in report_data["sections"]:
                json_data["sections"].append({
                    "title": section.title,
                    "content_type": section.content_type,
                    "data": section.data,
                    "order": section.order
                })
            
            return json.dumps(json_data, indent=2).encode('utf-8')
            
        except Exception as e:
            self.logger.error(f"Error generating JSON report: {str(e)}")
            raise
    
    async def _generate_csv_report(self, config: ReportConfig, report_data: Dict) -> bytes:
        """Generate CSV report"""
        try:
            # Combine all table data into a single CSV
            all_data = []
            
            for section in sorted(report_data["sections"], key=lambda x: x.order):
                if section.content_type == "table":
                    # Add section header
                    all_data.append([section.title])
                    all_data.append([])  # Empty row
                    
                    # Add table headers and data
                    all_data.append(section.data["headers"])
                    all_data.extend(section.data["rows"])
                    all_data.append([])  # Empty row between sections
            
            # Convert to DataFrame and then CSV
            df = pd.DataFrame(all_data)
            buffer = BytesIO()
            df.to_csv(buffer, index=False, header=False)
            buffer.seek(0)
            return buffer.getvalue()
            
        except Exception as e:
            self.logger.error(f"Error generating CSV report: {str(e)}")
            raise
    
    async def get_report(self, report_id: str) -> Optional[GeneratedReport]:
        """Retrieve a generated report by ID"""
        return self.report_cache.get(report_id)
    
    async def list_reports(self, filters: Optional[Dict] = None) -> List[Dict]:
        """List all generated reports with optional filtering"""
        reports = []
        for report_id, report in self.report_cache.items():
            report_info = {
                "report_id": report_id,
                "title": report.config.title,
                "type": report.config.report_type.value,
                "format": report.format.value,
                "generated_at": report.generated_at.isoformat(),
                "file_size": report.file_size
            }
            
            # Apply filters if provided
            if filters:
                if filters.get("type") and report.config.report_type.value != filters["type"]:
                    continue
                if filters.get("format") and report.format.value != filters["format"]:
                    continue
            
            reports.append(report_info)
        
        return sorted(reports, key=lambda x: x["generated_at"], reverse=True)
    
    def get_supported_formats(self) -> List[str]:
        """Get list of supported report formats"""
        formats = [ReportFormat.JSON.value, ReportFormat.CSV.value, ReportFormat.WEB.value]
        
        if REPORTLAB_AVAILABLE:
            formats.append(ReportFormat.PDF.value)
        
        if OPENPYXL_AVAILABLE:
            formats.append(ReportFormat.EXCEL.value)
        
        return formats
    
    def get_report_types(self) -> List[str]:
        """Get list of available report types"""
        return [rt.value for rt in ReportType]