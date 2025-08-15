"""
Export engine for generating PDF, Excel, and other format exports.
"""

import io
import json
import pandas as pd
from typing import Dict, List, Optional, Any, Union
from datetime import datetime
import logging
import base64
from pathlib import Path

try:
    from reportlab.lib.pagesizes import letter, A4
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Image, Table, TableStyle
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import inch
    from reportlab.lib import colors
    from reportlab.graphics.shapes import Drawing
    from reportlab.graphics.charts.linecharts import HorizontalLineChart
    from reportlab.graphics.charts.barcharts import VerticalBarChart
    from reportlab.graphics.charts.piecharts import Pie
    REPORTLAB_AVAILABLE = True
except ImportError:
    REPORTLAB_AVAILABLE = False

try:
    import openpyxl
    from openpyxl.chart import LineChart, BarChart, PieChart, Reference
    from openpyxl.drawing.image import Image as ExcelImage
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

from ..models.visualization_models import (
    ExportRequest, ExportFormat, VisualizationData, 
    DashboardLayout, PrintLayout, ChartType
)

logger = logging.getLogger(__name__)


class ExportEngine:
    """Advanced export engine for multiple formats."""
    
    def __init__(self):
        self.temp_dir = Path("temp/exports")
        self.temp_dir.mkdir(parents=True, exist_ok=True)
        
    async def export_visualization(
        self, 
        request: ExportRequest,
        visualizations: List[VisualizationData],
        dashboard: Optional[DashboardLayout] = None
    ) -> Dict[str, Any]:
        """Export visualizations in the requested format."""
        try:
            if request.format == ExportFormat.PDF:
                return await self._export_pdf(request, visualizations, dashboard)
            elif request.format == ExportFormat.EXCEL:
                return await self._export_excel(request, visualizations, dashboard)
            elif request.format == ExportFormat.CSV:
                return await self._export_csv(request, visualizations)
            elif request.format == ExportFormat.JSON:
                return await self._export_json(request, visualizations, dashboard)
            elif request.format in [ExportFormat.PNG, ExportFormat.SVG]:
                return await self._export_image(request, visualizations)
            else:
                return {
                    "success": False,
                    "error": f"Unsupported export format: {request.format}"
                }
                
        except Exception as e:
            logger.error(f"Export error: {str(e)}")
            return {
                "success": False,
                "error": f"Export failed: {str(e)}"
            }
    
    async def _export_pdf(
        self, 
        request: ExportRequest,
        visualizations: List[VisualizationData],
        dashboard: Optional[DashboardLayout] = None
    ) -> Dict[str, Any]:
        """Export to PDF format."""
        if not REPORTLAB_AVAILABLE:
            return {
                "success": False,
                "error": "ReportLab not available. Install with: pip install reportlab"
            }
        
        try:
            # Create PDF buffer
            buffer = io.BytesIO()
            
            # Create PDF document
            doc = SimpleDocTemplate(
                buffer,
                pagesize=A4,
                rightMargin=72,
                leftMargin=72,
                topMargin=72,
                bottomMargin=18
            )
            
            # Build content
            story = []
            styles = getSampleStyleSheet()
            
            # Title
            title = request.custom_title or (dashboard.name if dashboard else "Data Visualization Report")
            story.append(Paragraph(title, styles['Title']))
            story.append(Spacer(1, 12))
            
            # Description
            if request.custom_description or (dashboard and dashboard.description):
                desc = request.custom_description or dashboard.description
                story.append(Paragraph(desc, styles['Normal']))
                story.append(Spacer(1, 12))
            
            # Export metadata
            story.append(Paragraph(f"Generated on: {datetime.utcnow().strftime('%Y-%m-%d %H:%M:%S')} UTC", styles['Normal']))
            story.append(Spacer(1, 12))
            
            # Add visualizations
            for viz in visualizations:
                # Chart title
                story.append(Paragraph(viz.name, styles['Heading2']))
                story.append(Spacer(1, 6))
                
                # Chart description
                if viz.description:
                    story.append(Paragraph(viz.description, styles['Normal']))
                    story.append(Spacer(1, 6))
                
                # Add chart (simplified representation)
                chart_data = self._create_pdf_chart(viz)
                if chart_data:
                    story.append(chart_data)
                    story.append(Spacer(1, 12))
                
                # Add data table if requested
                if request.include_data and viz.data:
                    story.append(Paragraph("Data:", styles['Heading3']))
                    table_data = self._create_pdf_table(viz.data)
                    if table_data:
                        story.append(table_data)
                        story.append(Spacer(1, 12))
            
            # Build PDF
            doc.build(story)
            
            # Get PDF content
            pdf_content = buffer.getvalue()
            buffer.close()
            
            # Save to temp file
            filename = f"export_{datetime.utcnow().strftime('%Y%m%d_%H%M%S')}.pdf"
            filepath = self.temp_dir / filename
            
            with open(filepath, 'wb') as f:
                f.write(pdf_content)
            
            return {
                "success": True,
                "filename": filename,
                "filepath": str(filepath),
                "content_type": "application/pdf",
                "size": len(pdf_content)
            }
            
        except Exception as e:
            logger.error(f"PDF export error: {str(e)}")
            return {
                "success": False,
                "error": f"PDF export failed: {str(e)}"
            }
    
    async def _export_excel(
        self, 
        request: ExportRequest,
        visualizations: List[VisualizationData],
        dashboard: Optional[DashboardLayout] = None
    ) -> Dict[str, Any]:
        """Export to Excel format."""
        if not OPENPYXL_AVAILABLE:
            return {
                "success": False,
                "error": "openpyxl not available. Install with: pip install openpyxl"
            }
        
        try:
            # Create workbook
            wb = openpyxl.Workbook()
            
            # Remove default sheet
            wb.remove(wb.active)
            
            # Create summary sheet
            summary_sheet = wb.create_sheet("Summary")
            summary_sheet['A1'] = request.custom_title or "Data Visualization Export"
            summary_sheet['A2'] = f"Generated: {datetime.utcnow().strftime('%Y-%m-%d %H:%M:%S')} UTC"
            summary_sheet['A3'] = f"Number of visualizations: {len(visualizations)}"
            
            # Add visualization sheets
            for i, viz in enumerate(visualizations):
                sheet_name = f"Chart_{i+1}_{viz.name[:20]}"  # Limit sheet name length
                sheet = wb.create_sheet(sheet_name)
                
                # Add metadata
                sheet['A1'] = viz.name
                sheet['A2'] = viz.description or ""
                sheet['A3'] = f"Chart Type: {viz.chart_config.chart_type}"
                sheet['A4'] = f"Created: {viz.created_at.strftime('%Y-%m-%d %H:%M:%S')}"
                
                # Add data
                if viz.data:
                    df = pd.DataFrame(viz.data)
                    
                    # Write headers
                    for col_idx, column in enumerate(df.columns, 1):
                        sheet.cell(row=6, column=col_idx, value=column)
                    
                    # Write data
                    for row_idx, row in enumerate(df.itertuples(index=False), 7):
                        for col_idx, value in enumerate(row, 1):
                            sheet.cell(row=row_idx, column=col_idx, value=value)
                    
                    # Create chart if possible
                    try:
                        chart = self._create_excel_chart(viz, df)
                        if chart:
                            sheet.add_chart(chart, "H6")
                    except Exception as e:
                        logger.warning(f"Could not create Excel chart: {str(e)}")
            
            # Save to buffer
            buffer = io.BytesIO()
            wb.save(buffer)
            excel_content = buffer.getvalue()
            buffer.close()
            
            # Save to temp file
            filename = f"export_{datetime.utcnow().strftime('%Y%m%d_%H%M%S')}.xlsx"
            filepath = self.temp_dir / filename
            
            with open(filepath, 'wb') as f:
                f.write(excel_content)
            
            return {
                "success": True,
                "filename": filename,
                "filepath": str(filepath),
                "content_type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                "size": len(excel_content)
            }
            
        except Exception as e:
            logger.error(f"Excel export error: {str(e)}")
            return {
                "success": False,
                "error": f"Excel export failed: {str(e)}"
            }
    
    async def _export_csv(
        self, 
        request: ExportRequest,
        visualizations: List[VisualizationData]
    ) -> Dict[str, Any]:
        """Export to CSV format."""
        try:
            if len(visualizations) == 1:
                # Single CSV file
                viz = visualizations[0]
                df = pd.DataFrame(viz.data)
                
                filename = f"export_{datetime.utcnow().strftime('%Y%m%d_%H%M%S')}.csv"
                filepath = self.temp_dir / filename
                
                df.to_csv(filepath, index=False)
                
                return {
                    "success": True,
                    "filename": filename,
                    "filepath": str(filepath),
                    "content_type": "text/csv",
                    "size": filepath.stat().st_size
                }
            else:
                # Multiple CSV files in a zip
                import zipfile
                
                zip_filename = f"export_{datetime.utcnow().strftime('%Y%m%d_%H%M%S')}.zip"
                zip_filepath = self.temp_dir / zip_filename
                
                with zipfile.ZipFile(zip_filepath, 'w') as zipf:
                    for i, viz in enumerate(visualizations):
                        df = pd.DataFrame(viz.data)
                        csv_filename = f"chart_{i+1}_{viz.name[:20]}.csv"
                        
                        # Write CSV to memory
                        csv_buffer = io.StringIO()
                        df.to_csv(csv_buffer, index=False)
                        
                        # Add to zip
                        zipf.writestr(csv_filename, csv_buffer.getvalue())
                
                return {
                    "success": True,
                    "filename": zip_filename,
                    "filepath": str(zip_filepath),
                    "content_type": "application/zip",
                    "size": zip_filepath.stat().st_size
                }
                
        except Exception as e:
            logger.error(f"CSV export error: {str(e)}")
            return {
                "success": False,
                "error": f"CSV export failed: {str(e)}"
            }
    
    async def _export_json(
        self, 
        request: ExportRequest,
        visualizations: List[VisualizationData],
        dashboard: Optional[DashboardLayout] = None
    ) -> Dict[str, Any]:
        """Export to JSON format."""
        try:
            export_data = {
                "export_info": {
                    "generated_at": datetime.utcnow().isoformat(),
                    "format": "json",
                    "title": request.custom_title,
                    "description": request.custom_description,
                    "include_data": request.include_data,
                    "include_metadata": request.include_metadata
                },
                "visualizations": []
            }
            
            if dashboard:
                export_data["dashboard"] = {
                    "id": dashboard.id,
                    "name": dashboard.name,
                    "description": dashboard.description,
                    "layout": dashboard.layout,
                    "created_at": dashboard.created_at.isoformat(),
                    "updated_at": dashboard.updated_at.isoformat()
                }
            
            for viz in visualizations:
                viz_data = {
                    "id": viz.id,
                    "name": viz.name,
                    "description": viz.description,
                    "chart_config": viz.chart_config.dict(),
                    "created_at": viz.created_at.isoformat(),
                    "updated_at": viz.updated_at.isoformat()
                }
                
                if request.include_data:
                    viz_data["data"] = viz.data
                
                if request.include_metadata:
                    viz_data["metadata"] = viz.metadata
                
                export_data["visualizations"].append(viz_data)
            
            # Save to file
            filename = f"export_{datetime.utcnow().strftime('%Y%m%d_%H%M%S')}.json"
            filepath = self.temp_dir / filename
            
            with open(filepath, 'w') as f:
                json.dump(export_data, f, indent=2, default=str)
            
            return {
                "success": True,
                "filename": filename,
                "filepath": str(filepath),
                "content_type": "application/json",
                "size": filepath.stat().st_size
            }
            
        except Exception as e:
            logger.error(f"JSON export error: {str(e)}")
            return {
                "success": False,
                "error": f"JSON export failed: {str(e)}"
            }
    
    async def _export_image(
        self, 
        request: ExportRequest,
        visualizations: List[VisualizationData]
    ) -> Dict[str, Any]:
        """Export to image format (PNG/SVG)."""
        # This would require a headless browser or chart rendering library
        # For now, return a placeholder implementation
        return {
            "success": False,
            "error": "Image export not yet implemented. Requires headless browser setup."
        }
    
    def _create_pdf_chart(self, viz: VisualizationData) -> Optional[Any]:
        """Create a chart for PDF export."""
        if not viz.data:
            return None
        
        try:
            df = pd.DataFrame(viz.data)
            
            if viz.chart_config.chart_type == ChartType.BAR:
                # Create simple bar chart representation
                chart_data = []
                x_col = viz.chart_config.x_axis
                y_col = viz.chart_config.y_axis
                
                if isinstance(y_col, list):
                    y_col = y_col[0]
                
                if x_col in df.columns and y_col in df.columns:
                    for _, row in df.head(10).iterrows():  # Limit to 10 items
                        chart_data.append([str(row[x_col]), float(row[y_col])])
                
                if chart_data:
                    table = Table(chart_data)
                    table.setStyle(TableStyle([
                        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                        ('FONTSIZE', (0, 0), (-1, 0), 14),
                        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                        ('GRID', (0, 0), (-1, -1), 1, colors.black)
                    ]))
                    return table
            
        except Exception as e:
            logger.warning(f"Could not create PDF chart: {str(e)}")
        
        return None
    
    def _create_pdf_table(self, data: List[Dict[str, Any]]) -> Optional[Table]:
        """Create a data table for PDF export."""
        if not data:
            return None
        
        try:
            df = pd.DataFrame(data)
            
            # Limit rows and columns for PDF
            df = df.head(20)  # Max 20 rows
            if len(df.columns) > 6:  # Max 6 columns
                df = df.iloc[:, :6]
            
            # Create table data
            table_data = [list(df.columns)]
            for _, row in df.iterrows():
                table_data.append([str(val) for val in row.values])
            
            table = Table(table_data)
            table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 10),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('GRID', (0, 0), (-1, -1), 1, colors.black),
                ('FONTSIZE', (0, 1), (-1, -1), 8),
            ]))
            
            return table
            
        except Exception as e:
            logger.warning(f"Could not create PDF table: {str(e)}")
            return None
    
    def _create_excel_chart(self, viz: VisualizationData, df: pd.DataFrame) -> Optional[Any]:
        """Create a chart for Excel export."""
        try:
            x_col = viz.chart_config.x_axis
            y_col = viz.chart_config.y_axis
            
            if isinstance(y_col, list):
                y_col = y_col[0]
            
            if x_col not in df.columns or y_col not in df.columns:
                return None
            
            # Create chart based on type
            if viz.chart_config.chart_type == ChartType.LINE:
                chart = LineChart()
            elif viz.chart_config.chart_type == ChartType.BAR:
                chart = BarChart()
            elif viz.chart_config.chart_type == ChartType.PIE:
                chart = PieChart()
            else:
                return None
            
            chart.title = viz.name
            chart.style = 10
            chart.x_axis.title = x_col
            chart.y_axis.title = y_col
            
            # Add data (simplified)
            data = Reference(worksheet=None, min_col=2, min_row=7, max_row=min(7 + len(df), 27))
            cats = Reference(worksheet=None, min_col=1, min_row=7, max_row=min(7 + len(df), 27))
            
            chart.add_data(data, titles_from_data=True)
            chart.set_categories(cats)
            
            return chart
            
        except Exception as e:
            logger.warning(f"Could not create Excel chart: {str(e)}")
            return None